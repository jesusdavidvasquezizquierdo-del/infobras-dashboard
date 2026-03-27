#!/usr/bin/env python3
"""
cargar_datos.py — InfoBras DataSets → Supabase + SQLite

Lógica de fallback:
  1. Intenta el API de InfoBras en vivo
  2. Si falla → usa SQLite (caché de última descarga)
  3. Si SQLite vacío → carga desde archivos Excel/CSV descargados

Archivos soportados (carpeta supabase/):
  • DataSet-Obras-Publicas *.xlsx        (nacional, filtrar Lambayeque)
  • DataSet-Obras-Paralizadas *.xlsx     (nacional, filtrar Lambayeque)
  • DataSet-Obras-en-reconstruccion *.xlsx
  • DataSet-Asociaciones-Publico *.xlsx
  • Obras.xlsx                           (archivo adicional, ya filtrado)

Uso:
  python supabase/cargar_datos.py
Requisitos:
  pip install requests openpyxl
"""

import os, sys, time, glob, sqlite3, requests, zipfile
from datetime import datetime

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def load_env():
    for path in ['.env.local', '../.env.local',
                 os.path.join(os.path.dirname(__file__), '..', '.env.local')]:
        path = os.path.normpath(path)
        if os.path.exists(path):
            print(f"📄 Variables desde: {path}")
            with open(path) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        k, v = line.split('=', 1)
                        os.environ.setdefault(k.strip(), v.strip())
            return
    print("❌ No se encontró .env.local"); sys.exit(1)

load_env()

SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL', '').rstrip('/')
SERVICE_KEY  = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
if not SUPABASE_URL or not SERVICE_KEY:
    print("❌ Faltan variables en .env.local"); sys.exit(1)

INFOBRAS_BASE = "https://infobras.contraloria.gob.pe/InfobrasWeb"
INFOBRAS_API  = f"{INFOBRAS_BASE}/Busqueda/BusquedaSimple"   # nuevo endpoint 2025
INFOBRAS_HOME = f"{INFOBRAS_BASE}/Mapa/Index"                # página para obtener cookies
DEPARTAMENTO  = 14
PAGE_SIZE     = 100
DELAY         = 1.0

HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "es-PE,es;q=0.9",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://infobras.contraloria.gob.pe/InfobrasWeb/Mapa/Index",
}

# SQLite
_db_dirs = [
    os.path.join(os.path.dirname(__file__), '..', '..', 'infobras-scraper', 'data'),
    os.path.join(os.path.dirname(__file__), '..', 'data'),
    os.path.dirname(__file__), '.',
]
DB_PATH = next(
    (os.path.join(os.path.normpath(d), 'infobras_obras.db')
     for d in _db_dirs if os.path.exists(os.path.normpath(d))),
    'infobras_obras.db'
)
print(f"💾 Caché SQLite: {DB_PATH}")

# ══════════════════════════════════════════════════════════════════════════════
# PARSERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_fecha(v):
    if not v: return None
    s = str(v).strip()
    if s in ('', 'null', 'None', '-', 'nan'): return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try: return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except: pass
    return None

def parse_num(v):
    if v is None: return None
    try:
        s = str(v).replace(',', '').replace(' ', '').strip()
        return float(s) if s and s not in ('null','None','-','nan','') else None
    except: return None

def norm_estado(e):
    if not e: return None
    m = {'en ejecucion':'En Ejecución','en ejecución':'En Ejecución',
         'en ejecucion ':'En Ejecución','concluido':'Concluido',
         'finalizado':'Concluido','terminado':'Concluido',
         'paralizado':'Paralizada','paralizada':'Paralizada',
         'liquidado':'Liquidado','resuelto':'Resuelto','sin inicio':'Sin Inicio'}
    return m.get(str(e).lower().strip(), str(e).strip())

def es_lambayeque(v):
    if v is None: return True  # si no hay columna dept, asumir que ya está filtrado
    return 'LAMBAYEQUE' in str(v).upper()

# ══════════════════════════════════════════════════════════════════════════════
# SQLITE
# ══════════════════════════════════════════════════════════════════════════════

def sqlite_init():
    os.makedirs(os.path.dirname(os.path.abspath(DB_PATH)), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS obras (
        obra_id TEXT PRIMARY KEY, nombre_obra TEXT, municipio TEXT,
        entidad TEXT, estado TEXT, monto_aprobado REAL,
        avance_fisico REAL, fecha_inicio TEXT, fecha_fin TEXT,
        fuente TEXT, descargado_en TEXT DEFAULT (datetime('now')))""")
    conn.commit(); return conn

def sqlite_guardar(conn, obras, fuente='API'):
    for o in obras: o['fuente'] = fuente
    conn.executemany("""INSERT OR REPLACE INTO obras
        (obra_id,nombre_obra,municipio,entidad,estado,
         monto_aprobado,avance_fisico,fecha_inicio,fecha_fin,fuente)
        VALUES (:obra_id,:nombre_obra,:municipio,:entidad,:estado,
         :monto_aprobado,:avance_fisico,:fecha_inicio,:fecha_fin,:fuente)""", obras)
    conn.commit()

def sqlite_contar(conn):
    return conn.execute("SELECT COUNT(*) FROM obras").fetchone()[0]

def sqlite_cargar(conn):
    rows = conn.execute("SELECT * FROM obras").fetchall()
    cols = [d[0] for d in conn.execute("PRAGMA table_info(obras)").fetchall()]
    # PRAGMA devuelve: cid,name,type,... — extraemos los nombres
    col_names = [r[1] for r in conn.execute("PRAGMA table_info(obras)").fetchall()]
    rows2 = conn.execute("SELECT * FROM obras").fetchall()
    return [dict(zip(col_names, r)) for r in rows2]

# ══════════════════════════════════════════════════════════════════════════════
# FUENTE A: API InfoBras
# ══════════════════════════════════════════════════════════════════════════════

def _crear_sesion():
    """Crea sesión HTTP visitando la página principal para obtener cookies."""
    ses = requests.Session()
    ses.headers.update(HEADERS_HTTP)
    try:
        ses.get(INFOBRAS_HOME, timeout=15)  # obtiene cookies de sesión
        time.sleep(1)
    except Exception:
        pass
    return ses

def api_disponible():
    try:
        ses = _crear_sesion()
        r = ses.get(INFOBRAS_API, params={
            'nombre':'','departamento':14,'pageIndex':1,
            'pageSize':1,'orderBy':'en_ejecucion'},
            timeout=10)
        return r.status_code == 200 and r.text.strip().startswith('{')
    except: return False

def descargar_api():
    print("📡 Descargando desde API InfoBras (BusquedaSimple)...")
    ses = _crear_sesion()
    obras = []; page = 1
    while True:
        try:
            r = ses.get(INFOBRAS_API, params={
                'nombre':'','departamento':DEPARTAMENTO,
                'pageIndex':page,'pageSize':PAGE_SIZE,
                'orderBy':'en_ejecucion'}, timeout=30)
            if not r.text.strip().startswith('{'):
                print(f"  ⚠️  Respuesta no-JSON (pág {page}): {r.text[:100]}")
                break
            data    = r.json()
            results = data.get('Result', data.get('result', []))
            params  = data.get('Parameters', data.get('parameters', {}))
            total   = int(params.get('records', params.get('Records', len(results))))
            paginas = (total + PAGE_SIZE - 1) // PAGE_SIZE

            for raw in results:
                # BusquedaSimple puede tener nombres de campo distintos — intentamos ambos
                oid = str(raw.get('Codigo') or raw.get('CodigoObra') or raw.get('codigo') or '').strip()
                if not oid or oid == 'None': continue
                obras.append({
                    'obra_id':       oid,
                    'nombre_obra':   str(raw.get('NombreObra') or raw.get('Nombre') or '').strip() or None,
                    'municipio':     str(raw.get('Distrito') or raw.get('Provincia') or raw.get('Ubicacion') or '').strip() or None,
                    'entidad':       str(raw.get('EntidadNombre') or raw.get('Entidad') or '').strip() or None,
                    'estado':        norm_estado(raw.get('Estado') or raw.get('EstadoObra')),
                    'monto_aprobado':parse_num(raw.get('Monto') or raw.get('MontoAprobado')),
                    'avance_fisico': parse_num(raw.get('AvanceFisico') or raw.get('Avance')),
                    'fecha_inicio':  parse_fecha(raw.get('FechaInicioEjecucion') or raw.get('FechaInicio')),
                    'fecha_fin':     parse_fecha(raw.get('FechaFinalizacion') or raw.get('FechaFin')),
                })
            print(f"  ✓ Pág {page}/{paginas} — {len(obras)}/{total}")
            if len(obras) >= total or not results: break
            page += 1; time.sleep(DELAY)
        except Exception as e:
            print(f"  ❌ Error pág {page}: {e}"); break
    return obras

# ══════════════════════════════════════════════════════════════════════════════
# FUENTE B: Archivos Excel de InfoBras DataSets
# ══════════════════════════════════════════════════════════════════════════════

def detectar_fila_header(ws):
    """
    Los archivos de InfoBras tienen 3 filas de metadatos antes del header real.
    Detecta en qué fila está el encabezado buscando la fila con más columnas.
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        valores = [c for c in row if c is not None and str(c).strip()]
        if len(valores) >= 4:  # la fila del header tiene muchas columnas
            return i
        if i > 10: break
    return 0

def encontrar_col(encabezados, candidatos):
    enc_up = {str(h).strip().upper(): h for h in encabezados}
    for c in candidatos:
        if c.upper() in enc_up:
            return enc_up[c.upper()]
    return None

# Mapeo de columnas: nombre interno → posibles nombres en el Excel
COLS = {
    'obra_id':        ['CÓDIGO INFOBRAS', 'CODIGO INFOBRAS', 'INFOBRAS', 'Código INFOBRAS'],
    'nombre_obra':    ['NOMBRE DE OBRA', 'NOMBRE DE LA OBRA', 'NOMBRE OBRA'],
    'entidad':        ['ENTIDAD PÚBLICA', 'ENTIDAD PUBLICA', 'ENTIDAD'],
    'estado':         ['ESTADO DE EJECUCIÓN', 'ESTADO DE EJECUCION', 'ESTADO DE OBRA', 'ESTADO'],
    'municipio':      ['DISTRITO', 'MUNICIPIO'],
    'provincia':      ['PROVINCIA'],
    'departamento':   ['DEPARTAMENTO', 'REGION'],
    'monto_aprobado': ['MONTO APROBADO EN SOLES', 'MONTO EXPEDIENTE TÉCNICO',
                       'MONTO EXPEDIENTE TECNICO', 'MONTO VIABLE/APROBADO', 'MONTO'],
    'fecha_inicio':   ['FECHA DE INICIO DE OBRA', 'FECHA DE INICIO', 'FECHA INICIO'],
    'fecha_fin':      ['FECHA DE TÉRMINO CONTRACTUAL', 'FECHA DE FIN', 'FECHA FIN',
                       'FECHA DE TÉRMINO', 'FECHA TERMINO'],
    'avance_fisico':  ['AVANCE FÍSICO (%)', 'AVANCE FISICO', 'AVANCE (%)','AVANCE FÍSICO'],
}

def cargar_excel_infobras(ruta):
    """Carga un Excel de InfoBras, maneja las 3 filas de metadatos."""
    import openpyxl
    obras = []
    nombre = os.path.basename(ruta)

    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active

        # Leer todas las filas (paginado para no saturar RAM)
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            print(f"  ⚠️  {nombre}: archivo vacío")
            return []

        # Encontrar la fila del header (puede ser fila 1 o fila 4)
        header_idx = 0
        for i, row in enumerate(all_rows[:10]):
            vals = [c for c in row if c is not None and str(c).strip()]
            if len(vals) >= 4:
                header_idx = i
                break

        encabezados = [str(c or '').strip() for c in all_rows[header_idx]]

        # Mapear columnas
        col_map = {campo: encontrar_col(encabezados, cands)
                   for campo, cands in COLS.items()}

        # Verificar que encontramos la columna esencial
        if not col_map.get('obra_id'):
            print(f"  ⚠️  {nombre}: no se encontró columna 'Código INFOBRAS'")
            print(f"      Columnas disponibles: {encabezados[:10]}")
            return []

        print(f"  📊 {nombre}: {len(all_rows)-header_idx-1} filas, "
              f"header en fila {header_idx+1}")

        # Procesar filas de datos
        filtradas = 0
        for row in all_rows[header_idx+1:]:
            fila = dict(zip(encabezados, row))

            def g(campo):
                col = col_map.get(campo)
                return fila.get(col) if col else None

            # Filtrar por Lambayeque
            dept = g('departamento')
            if dept is not None and not es_lambayeque(dept):
                filtradas += 1
                continue

            oid = str(g('obra_id') or '').strip()
            if not oid or oid in ('nan', 'None', '', 'None'):
                continue

            municipio = g('municipio') or g('provincia')

            obras.append({
                'obra_id':        oid,
                'nombre_obra':    str(g('nombre_obra') or '').strip() or None,
                'municipio':      str(municipio or '').strip() or None,
                'entidad':        str(g('entidad') or '').strip() or None,
                'estado':         norm_estado(g('estado')),
                'monto_aprobado': parse_num(g('monto_aprobado')),
                'avance_fisico':  parse_num(g('avance_fisico')),
                'fecha_inicio':   parse_fecha(g('fecha_inicio')),
                'fecha_fin':      parse_fecha(g('fecha_fin')),
            })

        print(f"  ✅ {nombre}: {len(obras)} obras de Lambayeque "
              f"({filtradas} de otros departamentos descartadas)")
        return obras

    except Exception as e:
        print(f"  ❌ Error leyendo {nombre}: {e}")
        return []

def cargar_desde_archivos():
    """Busca y carga todos los Excel de InfoBras en la carpeta supabase/."""
    carpeta = os.path.dirname(os.path.abspath(__file__))
    archivos = glob.glob(os.path.join(carpeta, '*.xlsx')) + \
               glob.glob(os.path.join(carpeta, '*.xls')) + \
               glob.glob(os.path.join(carpeta, '*.csv'))

    if not archivos:
        return []

    print(f"\n📂 Archivos encontrados en {carpeta}:")
    for a in archivos:
        print(f"   • {os.path.basename(a)}")

    todas = {}
    for ruta in archivos:
        ext = os.path.splitext(ruta)[1].lower()
        if ext in ('.xlsx', '.xls'):
            obras = cargar_excel_infobras(ruta)
        else:
            continue  # CSV support puede agregarse si hace falta
        for o in obras:
            # Merge: si ya existe la obra, enriquecer datos faltantes
            oid = o['obra_id']
            if oid not in todas:
                todas[oid] = o
            else:
                existing = todas[oid]
                for k, v in o.items():
                    if v is not None and existing.get(k) is None:
                        existing[k] = v

    result = list(todas.values())
    print(f"\n📦 Total obras únicas de Lambayeque: {len(result)}")
    return result

# ══════════════════════════════════════════════════════════════════════════════
# SUPABASE
# ══════════════════════════════════════════════════════════════════════════════

SUP_HDR = {
    'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates',
}

def subir_supabase(obras, batch=300):
    url = f"{SUPABASE_URL}/rest/v1/obras"
    total = 0; errores = 0
    # Campos que acepta Supabase (sin columnas generadas ni extras)
    campos_ok = {'obra_id','nombre_obra','municipio','entidad','estado',
                 'monto_aprobado','avance_fisico','fecha_inicio','fecha_fin'}
    for i in range(0, len(obras), batch):
        lote = [{k:v for k,v in o.items() if k in campos_ok}
                for o in obras[i:i+batch]]
        try:
            r = requests.post(url, headers=SUP_HDR, json=lote, timeout=60)
            if r.status_code in (200, 201):
                total += len(lote)
                print(f"  ✓ Lote {i//batch+1}: {total}/{len(obras)} obras")
            else:
                print(f"  ❌ HTTP {r.status_code}: {r.text[:200]}")
                errores += len(lote)
        except Exception as e:
            print(f"  ❌ {e}"); errores += len(lote)
        time.sleep(0.3)
    return total, errores

# ══════════════════════════════════════════════════════════════════════════════
# MAIN — fallback automático
# ══════════════════════════════════════════════════════════════════════════════

def main():
    t0 = datetime.now()
    print("\n" + "="*60)
    print("  InfoBras → Supabase  |  Lambayeque")
    print("="*60)

    # Verificar Supabase
    print("\n🔍 Verificando Supabase...")
    try:
        r = requests.get(f"{SUPABASE_URL}/rest/v1/obras",
                         headers={**SUP_HDR,'Prefer':'count=exact'},
                         params={'select':'count'}, timeout=10)
        cr = r.headers.get('Content-Range','?/?')
        print(f"✅ Supabase OK — obras actuales: {cr.split('/')[-1]}")
    except Exception as e:
        print(f"⚠️  Supabase: {e}")

    conn   = sqlite_init()
    obras  = []
    fuente = None

    # ── Fuente 1: API InfoBras ───────────────────────────────────────────────
    print("\n🌐 Verificando API InfoBras...")
    if api_disponible():
        print("✅ API disponible — descargando...")
        obras = descargar_api()
        if obras:
            fuente = 'API_INFOBRAS'
            sqlite_guardar(conn, obras, 'API')
    else:
        print("⚠️  API no disponible (sitio en mantenimiento o caído)")

    # ── Fuente 2: SQLite caché ───────────────────────────────────────────────
    if not obras:
        n = sqlite_contar(conn)
        if n > 0:
            print(f"\n📦 Usando caché SQLite: {n} obras de última descarga exitosa")
            obras  = sqlite_cargar(conn)
            fuente = 'CACHE_SQLITE'
        else:
            print("\n📦 SQLite vacío — sin caché")

    # ── Fuente 3: Archivos Excel/CSV ─────────────────────────────────────────
    if not obras:
        print("\n📂 Cargando desde archivos Excel descargados de InfoBras...")
        obras = cargar_desde_archivos()
        if obras:
            fuente = 'EXCEL_LOCAL'
            print(f"\n💾 Guardando {len(obras)} obras en SQLite (caché)...")
            sqlite_guardar(conn, obras, 'EXCEL')
        else:
            print("\n❌ Sin datos. Opciones:")
            print("   1. Espera a que InfoBras vuelva y re-ejecuta este script")
            print("   2. Descarga datasets desde:")
            print("      https://infobras.contraloria.gob.pe/InfobrasWeb/DataSets")
            print(f"   3. Copia los archivos .xlsx a: {os.path.dirname(__file__)}")
            conn.close(); sys.exit(1)

    conn.close()

    # ── Subir a Supabase ─────────────────────────────────────────────────────
    print(f"\n☁️  Subiendo {len(obras)} obras a Supabase (fuente: {fuente})...")
    subidas, errores = subir_supabase(obras)

    seg = int((datetime.now() - t0).total_seconds())
    print("\n" + "="*60)
    print(f"  🎉 COMPLETADO en {seg}s")
    print(f"  📊 Obras procesadas : {len(obras)}")
    print(f"  ☁️  Subidas OK       : {subidas}  |  Errores: {errores}")
    print(f"  🔀 Fuente           : {fuente}")
    print(f"  💾 SQLite caché     : {DB_PATH}")
    print("="*60)
    print("\n👉 Abre: http://localhost:3000")

if __name__ == '__main__':
    main()
